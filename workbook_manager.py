#!/usr/bin/env python3
"""
workbook_manager.py — Monthly workbook rollover logic.

Handles:
- report_date resolution (yesterday in Vietnam TZ, --report-date override, or --run-date simulation)
- month_key = YYYY-MM from report_date
- Pre-created monthly workbook path: reports/51la_YYYY-MM.xlsx
- Optional auto-create mode (behind explicit opt-in)
- Optional copy to 51la_current.xlsx
"""

import os
import shutil
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional, Tuple

import openpyxl

from scraper import COLUMNS, DATE_COLUMN, METRIC_COLUMNS, copy_row_style, normalize_excel_date


# ==================== CUSTOM EXCEPTIONS ====================

class MonthlyWorkbookMissingError(FileNotFoundError):
    """Raised when a pre-created monthly workbook does not exist."""
    pass


class DateRowNotFoundError(ValueError):
    """Raised when the report date row is not found in the monthly workbook."""
    pass


# ==================== REPORT DATE ====================

def get_report_date(override_date: Optional[str] = None, run_date: Optional[date] = None) -> date:
    """
    Return report_date = yesterday in Asia/Ho_Chi_Minh timezone.

    Modes:
      --report-date YYYY-MM-DD  : direct override of report_date (ignores run date entirely)
      --run-date YYYY-MM-DD      : simulate current run date; report_date = run_date - 1 day
      neither passed             : report_date = yesterday of real current time

    --report-date and --run-date are mutually exclusive.
    """
    if override_date:
        return datetime.strptime(override_date, "%Y-%m-%d").date()

    from zoneinfo import ZoneInfo
    vn_tz = ZoneInfo("Asia/Ho_Chi_Minh")

    if run_date:
        # Simulate current time = run_date, so yesterday = run_date - 1 day
        computed = datetime.combine(run_date, datetime.min.time())
        yesterday = computed - timedelta(days=1)
        return yesterday.date()

    # Real run: use actual current time
    now_vn = datetime.now(vn_tz)
    yesterday = now_vn - timedelta(days=1)
    return yesterday.date()


def get_month_key(report_date: date) -> str:
    """Return YYYY-MM string for a given date."""
    return f"{report_date.year}-{report_date.month:02d}"


def get_month_workbook_path(report_date: date, reports_dir: str) -> Path:
    """Return reports/51la_YYYY-MM.xlsx for a given report_date."""
    month_key = get_month_key(report_date)
    return Path(reports_dir) / f"51la_{month_key}.xlsx"


# ==================== TEMPLATE ====================

def ensure_template(template_path: Path, source_workbook_path: Path) -> Tuple[Path, bool]:
    """
    Ensure template exists at template_path.
    If missing: create from source_workbook_path with F/G/H/I cleared.
    Returns (template_path, created_bool).
    """
    if template_path.exists():
        return template_path, False

    print(f"[INFO] Template not found at {template_path}, creating from {source_workbook_path}...")
    try:
        wb = openpyxl.load_workbook(source_workbook_path)
    except Exception as e:
        raise RuntimeError(f"Cannot open source workbook {source_workbook_path}: {e}")

    ws = wb.active

    # Clear F/G/H/I for all data rows (keep header)
    for row in range(2, ws.max_row + 1):
        date_val = normalize_excel_date(ws[f"{DATE_COLUMN}{row}"].value)
        if date_val is None:
            continue
        for col in COLUMNS:
            ws[f"{col}{row}"] = None

    template_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(template_path)
    wb.close()
    print(f"[INFO] Template created: {template_path}")
    return template_path, True


# ==================== PRE-CREATED MONTHLY WORKBOOK ====================

def resolve_precreated_monthly_workbook(
    report_date: date,
    reports_dir: str,
    pattern: str = "51la_{YYYY-MM}.xlsx",
) -> Path:
    """
    Resolve path to a pre-created monthly workbook.

    Raises MonthlyWorkbookMissingError if the file does not exist.
    Does NOT create the file.
    """
    month_key = get_month_key(report_date)
    filename = pattern.replace("{YYYY-MM}", month_key)
    path = Path(reports_dir) / filename

    if not path.exists():
        raise MonthlyWorkbookMissingError(
            f"Monthly workbook not found: {path}. Please create this file before running."
        )

    return path


def find_date_row_in_workbook(workbook_path: Path, report_date: date, sheet_name: str = "Sheet1") -> int:
    """
    Find the row in the workbook matching report_date (by column B).

    Raises DateRowNotFoundError if the row is not found.
    Does NOT insert or create rows.
    """
    wb = openpyxl.load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise DateRowNotFoundError(f"Sheet '{sheet_name}' not found in {workbook_path}")

    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        normalized = normalize_excel_date(ws[f"{DATE_COLUMN}{row}"].value)
        if normalized and normalized.year == report_date.year and normalized.month == report_date.month and normalized.day == report_date.day:
            wb.close()
            return row

    wb.close()
    raise DateRowNotFoundError(
        f"Date row {report_date.strftime('%Y-%m-%d')} not found in {workbook_path}. "
        "Please check the prepared monthly workbook."
    )


# ==================== AUTO-CREATE MONTHLY WORKBOOK ====================

def _get_days_in_month(year: int, month: int) -> int:
    """Return number of days in the given month."""
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    return (next_month - date(year, month, 1)).days


def create_month_workbook_from_template(
    template_path: Path,
    workbook_path: Path,
    report_date: date,
) -> Path:
    """
    Copy template to monthly workbook path and initialize all date rows.
    Fills column B with all dates in the report_date's month.
    Clears F/G/H/I for all rows.
    """
    if workbook_path.exists():
        return workbook_path

    print(f"[INFO] Creating monthly workbook from template: {workbook_path}")
    shutil.copy2(template_path, workbook_path)

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active

    # Unmerge any merged cells so we can write to them
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    # Find last row with data in template (to copy style from)
    last_data_row = 1
    for row in range(2, ws.max_row + 1):
        if ws[f"{DATE_COLUMN}{row}"].value is not None:
            last_data_row = row

    # Initialize all days of the month
    year, month = report_date.year, report_date.month
    days_in_month = _get_days_in_month(year, month)

    # Check if template already has rows beyond header
    existing_rows = ws.max_row

    for day in range(1, days_in_month + 1):
        target_date = datetime(year, month, day, 0, 0, 0)
        row_num = day + 1  # row 2 = day 1, etc.

        # Extend sheet if needed
        if row_num > existing_rows:
            ws.insert_rows(row_num)
            copy_row_style(ws, last_data_row, row_num)

        ws[f"{DATE_COLUMN}{row_num}"] = target_date

        # Clear metric columns
        for col in COLUMNS:
            ws[f"{col}{row_num}"] = None

    wb.save(workbook_path)
    wb.close()
    print(f"[INFO] Monthly workbook initialized: {workbook_path} ({days_in_month} days)")
    return workbook_path


# ==================== RESOLVE WORKBOOK ====================

def get_workbook_mode(cfg, args) -> str:
    """
    Determine which workbook mode would be used, without resolving the actual path.

    This is stable across runs (does not depend on report_date), so it can be
    computed once at startup for safety-gate checks. Use resolve_workbook_for_report
    to also get the path (which DOES depend on report_date and must be re-resolved
    per scheduled run).
    """
    explicit_excel_path = getattr(args, "excel_path", None)
    use_precreated = (
        getattr(args, "precreated_monthly", False)
        or getattr(args, "monthly_workbook", False)
        or getattr(cfg, "use_precreated_monthly_files", False)
    )
    use_auto_create = getattr(args, "auto_create_monthly", False) or getattr(cfg, "auto_create_monthly_workbook", False)
    use_legacy = getattr(args, "legacy_excel", False)

    if explicit_excel_path and not use_precreated and not use_auto_create and not use_legacy:
        return "legacy"
    if use_legacy:
        return "legacy"
    if use_precreated:
        return "precreated"
    if use_auto_create:
        return "auto_create"
    return "legacy"


def resolve_workbook_for_report(
    cfg,
    args,
    report_date: date,
) -> Tuple[Path, str, bool]:
    """
    Resolve the workbook path to use for the current report.

    Returns (workbook_path, mode, was_created).

    mode: "precreated" | "auto_create" | "legacy"
    was_created: True only when auto-create created a new file
    """
    use_precreated = getattr(args, "precreated_monthly", False) or getattr(args, "monthly_workbook", False)
    use_auto_create = getattr(args, "auto_create_monthly", False)
    use_legacy = getattr(args, "legacy_excel", False)
    env_precreated = getattr(cfg, "use_precreated_monthly_files", False)
    env_auto_create = getattr(cfg, "auto_create_monthly_workbook", False)

    # Force legacy if EXCEL_PATH is explicitly set via --excel-path and not using any monthly mode
    explicit_excel_path = getattr(args, "excel_path", None)
    if explicit_excel_path and not use_precreated and not use_auto_create and not use_legacy:
        return Path(explicit_excel_path), "legacy", False

    if use_legacy:
        return Path(cfg.excel_path), "legacy", False

    # Pre-created monthly mode (CLI flag or env default)
    if use_precreated or env_precreated:
        reports_dir = getattr(cfg, "reports_dir", "reports")
        pattern = getattr(cfg, "report_file_pattern", "51la_{YYYY-MM}.xlsx")
        path = resolve_precreated_monthly_workbook(report_date, reports_dir, pattern)
        return path, "precreated", False

    # Auto-create monthly mode (only behind explicit opt-in)
    if use_auto_create or env_auto_create:
        print("[WARN] Auto-create monthly workbook may not preserve custom Excel layout. Pre-created monthly files are recommended.")
        reports_dir = getattr(cfg, "reports_dir", "reports")
        reports_path = Path(reports_dir)
        reports_path.mkdir(parents=True, exist_ok=True)

        workbook_path = get_month_workbook_path(report_date, reports_dir)

        if workbook_path.exists():
            return workbook_path, "auto_create", False

        # Create new monthly workbook
        template_path = Path(getattr(cfg, "template_path", "templates/51la_template.xlsx"))
        source_workbook = Path(cfg.excel_path)

        try:
            template_path, _ = ensure_template(template_path, source_workbook)
        except RuntimeError as e:
            raise RuntimeError(f"Cannot create monthly workbook: {e}")

        created_path = create_month_workbook_from_template(template_path, workbook_path, report_date)
        return created_path, "auto_create", True

    # Default: legacy mode
    return Path(cfg.excel_path), "legacy", False


# ==================== MISMATCH GUARD ====================

def validate_workbook_month_match(report_date: date, workbook_path: Path) -> None:
    """
    Raise ValueError if the workbook filename's YYYY-MM does not match report_date's month.

    For pre-created / auto-create monthly mode, the workbook filename encodes the
    month it covers. If the resolved workbook is for a different month than
    report_date, that means report_date or workbook resolution is wrong — block
    the write to avoid corrupting the wrong file.
    """
    expected_month_key = get_month_key(report_date)
    if expected_month_key not in workbook_path.name:
        raise ValueError(
            f"BLOCKED: Workbook/report_date mismatch.\n"
            f"  Report date:  {report_date.strftime('%Y-%m-%d')}\n"
            f"  Workbook:     {workbook_path}\n"
            f"  Expected month key: {expected_month_key}"
        )


# ==================== CURRENT COPY ====================

def copy_to_current(monthly_path: Path, current_path: Path) -> None:
    """Copy monthly workbook to 51la_current.xlsx if enabled."""
    if not monthly_path.exists():
        print(f"[WARN] Monthly workbook not found for current copy: {monthly_path}")
        return
    try:
        shutil.copy2(monthly_path, current_path)
        print(f"[INFO] Current copy updated: {current_path}")
    except Exception as e:
        print(f"[WARN] Could not update current copy: {e}")