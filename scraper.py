#!/usr/bin/env python3
"""
scraper.py — Scrape UV/PV from 51.la, write to Excel, capture screenshots.
Preserves existing selectors, timeout, password login, and fallback behavior.
Fixes: timezone (ZoneInfo), return type (always ScrapeResult), Excel COM fallback,
Excel auto-row-creation when date is missing.
"""

import asyncio
import os
import re
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from typing import List, Literal, Optional, Tuple
from zoneinfo import ZoneInfo

import openpyxl
from playwright.async_api import async_playwright, Page, TimeoutError as PlaywrightTimeout

# Excel screenshot — Windows only with pywin32
try:
    from PIL import Image, ImageDraw, ImageFont
    import win32com.client
    from PIL import ImageGrab
    HAS_COM = True
except ImportError:
    HAS_COM = False

# ==================== CONFIG ====================
LINKS: List[str] = []
PASSWORD = ""

UV_SELECTOR = "#app > div > div.main.clearfix > div.mt-54.content > div.report > div:nth-child(1) > div > div > div.report-overview > div.overview.overview.overview-inactive > div.overview-top-right > div.expandable-box.layout-inactive > div.daily > div:nth-child(4) > dd:nth-child(3)"
PV_SELECTOR = "#app > div > div.main.clearfix > div.mt-54.content > div.report > div:nth-child(1) > div > div > div.report-overview > div.overview.overview.overview-inactive > div.overview-top-right > div.expandable-box.layout-inactive > div.daily > div:nth-child(1) > dd:nth-child(3)"

COLUMNS = ["F", "G", "H", "I"]
METRIC_COLUMNS = {"F", "G", "H", "I"}
DATE_COLUMN = "B"

PAGE_TIMEOUT = 60000
ELEMENT_TIMEOUT = 15000
SCREENSHOT_DIR = "screenshots"

# ==================== DATA CLASSES ====================

# Page-level status after scraping
# PASS: numeric UV found successfully
# NO_DATA: selector found data but it was confirmed zero (e.g. 51.la shows 0 for that date)
# SCRAPE_ERROR: selector failed or couldn't extract numeric data
ScrapeStatus = Literal["PASS", "NO_DATA", "SCRAPE_ERROR"]


@dataclass
class ScrapeResult:
    index: int
    uv: int
    pv: int
    screenshot_path: str
    status: ScrapeStatus          # replaces 'success' bool
    error: Optional[str] = None   # populated on SCRAPE_ERROR only


@dataclass
class ScrapeReport:
    uv_results: List[int]
    pv_results: List[int]
    screenshot_paths: List[str]
    statuses: List[ScrapeStatus]   # per-page status
    failed_indexes: List[int]      # indexes where status == SCRAPE_ERROR
    data_date: Optional[datetime] = None  # the date this data belongs to


# ==================== TIMEZONE ====================

VIETNAM_TZ = ZoneInfo("Asia/Ho_Chi_Minh")

# Override for monthly workbook mode (set by main.py)
_report_date_override: Optional[datetime] = None


def get_data_date() -> datetime:
    """
    Return the date for which we want to scrape 51.la data.
    
    The 51.la "昨日" (yesterday) metric is finalized only after ~10 AM Vietnam time.
    If we run before 10 AM, "yesterday" data is still incomplete.
    To get complete data, we must scrape the day before "yesterday".
    
    This function computes the correct data_date based on the current hour.
    """
    now_vn = datetime.now(VIETNAM_TZ)
    yesterday = (now_vn - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # If running before 10 AM Vietnam time, yesterday's data is incomplete
    # (51.la finalizes previous day data around 10 AM)
    if now_vn.hour < 10:
        # Get the day before yesterday for complete data
        data_date = (now_vn - timedelta(days=2)).replace(hour=0, minute=0, second=0, microsecond=0)
        print(f"[INFO] Running before 10 AM — scraping {data_date.strftime('%Y-%m-%d')} (day before yesterday) for complete data")
    else:
        data_date = yesterday
        print(f"[INFO] Running after 10 AM — scraping {data_date.strftime('%Y-%m-%d')} (yesterday) for complete data")
    
    return data_date


def set_report_date_override(report_date: Optional[date]) -> None:
    """Set the report date override for monthly workbook mode. Pass date object."""
    global _report_date_override
    if report_date is None:
        _report_date_override = None
    else:
        _report_date_override = datetime.combine(report_date, datetime.min.time())


def get_yesterday_vietnam() -> datetime:
    """Return yesterday's date in Vietnam TZ. Respects report_date_override for monthly mode."""
    if _report_date_override is not None:
        return _report_date_override
    now_vn = datetime.now(VIETNAM_TZ)
    yesterday = (now_vn - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    return yesterday.replace(tzinfo=None)


# ==================== SCRAPING HELPERS ====================

def _extract_numeric(text: str) -> Optional[int]:
    """
    Extract first non-negative integer from text.
    Handles comma-separated numbers like '75,812'.
    Returns None if no valid non-negative integer found.
    Returns 0 when text is exactly '0' (valid zero from 51.la).
    """
    text = text.strip()
    if not text:
        return None
    cleaned = re.sub(r'[\s,]', '', text)
    if cleaned.isdigit() and int(cleaned) >= 0:
        return int(cleaned)
    return None


async def _save_debug_artifacts(page: Page, index: int, error_msg: str):
    """
    Save sanitized debug artifacts when scraping fails.
    Screenshot + HTML dump. No secrets.
    """
    n = index + 1
    debug_dir = os.path.join(SCREENSHOT_DIR, "debug")
    os.makedirs(debug_dir, exist_ok=True)

    # Screenshot
    ss_path = os.path.join(debug_dir, f"failed_{n}.png")
    try:
        await page.screenshot(path=ss_path, full_page=False)
        print(f"[{n}] Debug screenshot: {ss_path}")
    except Exception as e:
        print(f"[{n}] Debug screenshot failed: {e}")

    # HTML dump — redact URLs and password inputs
    html_path = os.path.join(debug_dir, f"failed_{n}.html")
    try:
        content = await page.content()
        content = re.sub(r'https?://[^\s"\'<>]+(\?[^"\s\'<>]*)', '[URL_REDACTED]', content)
        content = re.sub(r'<input[^>]*type=["\']?password["\']?[^>]*>', '[PASSWORD_INPUT]', content, flags=re.IGNORECASE)
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"[{n}] Debug HTML saved: {html_path}")
    except Exception as e:
        print(f"[{n}] Debug HTML failed: {e}")


# ==================== MAIN SCRAPE FUNCTION ====================

async def scrape_uv_from_link(
    page: Page,
    link: str,
    index: int,
    screenshot_path: str
) -> ScrapeResult:
    """
    Access a 51.la stats page and extract yesterday's UV and PV.
    Always returns a ScrapeResult with explicit status:
      PASS        — numeric UV found and extracted
      NO_DATA     — page loaded but returned 0 (confirmed zero from 51.la)
      SCRAPE_ERROR — selector failed or page didn't load
    On SCRAPE_ERROR, saves debug artifacts (screenshot + sanitized HTML).
    """
    n = index + 1
    total = len(LINKS)

    try:
        print(f"[{n}/{total}] Opening: {link[:40]}...")
        await page.goto(link, timeout=PAGE_TIMEOUT)
        await asyncio.sleep(2)

        # Password gate
        password_input = await page.query_selector('input[type="password"]')
        if password_input:
            print(f"[{n}/{total}] Entering password...")
            await password_input.fill(PASSWORD)
            await password_input.press("Enter")
            await asyncio.sleep(5)
        else:
            print(f"[{n}/{total}] No password required, waiting...")
            await asyncio.sleep(3)

        # ---- UV extraction ----
        # Track what we found at each selector
        uv_text_found: Optional[str] = None  # raw text from selector

        # Primary selector: full path to UV value (column 昨日 = yesterday)
        # Fallback: find the UV row by its dt label, then get the 2nd dd (昨日 column)
        uv_selectors = [
            UV_SELECTOR,
            "dd:nth-child(3)",
            ".daily-right-item:has(dt:text('访客数(UV)')) dd:nth-child(2)",
            ".daily dd:nth-child(3)",
            "[class*='daily'] dd:nth-child(3)",
        ]

        for selector in uv_selectors:
            try:
                print(f"[{n}/{total}] Trying UV selector: {selector[:50]}...")
                await page.wait_for_selector(selector, timeout=5000)
                elem = await page.query_selector(selector)
                if elem:
                    text = await elem.inner_text()
                    print(f"[{n}/{total}] Got text: '{text}'")
                    uv_text_found = text
                    # Try to extract numeric value (0 is valid data from 51.la)
                    uv = _extract_numeric(text)
                    if uv is not None:
                        # Got a valid non-negative integer
                        uv_value = uv
                        break
                    # uv is None means text wasn't numeric
                    # continue to next selector
            except Exception:
                continue

        # Determine UV result
        if uv_text_found is None:
            # No selector found anything at all
            err = "No UV selector matched; page may not have loaded"
            print(f"[{n}/{total}] SCRAPE_ERROR: {err}")
            await _save_debug_artifacts(page, index, err)
            return ScrapeResult(index=index, uv=0, pv=0, screenshot_path=screenshot_path, status="SCRAPE_ERROR", error=err)

        # uv_text_found has the text; check if we got a number
        uv = _extract_numeric(uv_text_found)
        if uv is None:
            # Selector matched but returned non-numeric text like "昨日"
            err = f"No numeric UV found; selector returned label text: {uv_text_found!r}"
            print(f"[{n}/{total}] SCRAPE_ERROR: {err}")
            await _save_debug_artifacts(page, index, err)
            return ScrapeResult(index=index, uv=0, pv=0, screenshot_path=screenshot_path, status="SCRAPE_ERROR", error=err)

        # We got a numeric value — uv >= 0 is valid data from 51.la
        # uv > 0 = PASS, uv == 0 = NO_DATA
        uv_value = uv

        # ---- PV extraction ----
        pv_value = 0
        pv_selectors = [
            ".daily-right-item:has(dt:text('浏览量(PV)')) dd:nth-child(3)",
            ".daily dd:nth-child(3)",
            ".daily-right-item:has(dt:text('访客数(UV)')) dd:nth-child(3)",
            "[class*='daily'] dd:nth-child(3)",
        ]

        for selector in pv_selectors:
            try:
                await page.wait_for_selector(selector, timeout=3000)
                elem = await page.query_selector(selector)
                if elem:
                    text = await elem.inner_text()
                    pv = _extract_numeric(text)
                    if pv is not None:
                        pv_value = pv
                        print(f"[{n}/{total}] Got PV = {pv_value:,}")
                        break
            except Exception:
                continue

        # ---- Result ----
        if uv_value == 0:
            # Confirmed zero from 51.la — not a scrape error
            print(f"[{n}/{total}] NO_DATA: UV = 0 (confirmed zero from 51.la)")
            status: ScrapeStatus = "NO_DATA"
        else:
            print(f"[{n}/{total}] PASS: UV = {uv_value:,} | PV = {pv_value:,}")
            status = "PASS"

        # Screenshot
        print(f"[{n}/{total}] Taking screenshot...")
        try:
            await page.screenshot(path=screenshot_path, full_page=False)
            print(f"[{n}/{total}] Screenshot saved: {screenshot_path}")
        except Exception as e:
            print(f"[{n}/{total}] Screenshot error: {str(e)[:50]}...")

        return ScrapeResult(index=index, uv=uv_value, pv=pv_value, screenshot_path=screenshot_path, status=status)

    except PlaywrightTimeout:
        err = "Timeout waiting for page"
        print(f"[{n}/{total}] SCRAPE_ERROR: {err}")
        await _save_debug_artifacts(page, index, err)
        return ScrapeResult(index=index, uv=0, pv=0, screenshot_path=screenshot_path, status="SCRAPE_ERROR", error=err)
    except Exception as e:
        err = str(e)[:80]
        print(f"[{n}/{total}] SCRAPE_ERROR: {err}")
        await _save_debug_artifacts(page, index, err)
        return ScrapeResult(index=index, uv=0, pv=0, screenshot_path=screenshot_path, status="SCRAPE_ERROR", error=err)


async def scrape_all_uv(data_date: Optional[datetime] = None) -> ScrapeReport:
    """
    Run 4 scrape tasks concurrently with Semaphore(2) limit.
    Returns ScrapeReport with UV, PV, screenshot paths, per-page statuses, and failed indexes.
    """
    uv_results = [0] * len(LINKS)
    pv_results = [0] * len(LINKS)
    screenshot_paths = [""] * len(LINKS)
    statuses: List[ScrapeStatus] = ["SCRAPE_ERROR"] * len(LINKS)
    failed_indexes: List[int] = []

    os.makedirs(SCREENSHOT_DIR, exist_ok=True)

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=True)
        semaphore = asyncio.Semaphore(2)

        async def scrape_with_semaphore(link: str, index: int) -> ScrapeResult:
            async with semaphore:
                context = await browser.new_context()
                page = await context.new_page()
                screenshot_path = os.path.join(SCREENSHOT_DIR, f"screenshot_{index + 1}.png")
                try:
                    result = await scrape_uv_from_link(page, link, index, screenshot_path)
                    uv_results[index] = result.uv
                    pv_results[index] = result.pv
                    screenshot_paths[index] = result.screenshot_path
                    statuses[index] = result.status
                    if result.status == "SCRAPE_ERROR":
                        failed_indexes.append(index)
                    return result
                finally:
                    await context.close()

        tasks = [scrape_with_semaphore(link, i) for i, link in enumerate(LINKS)]
        results = await asyncio.gather(*tasks)
        await browser.close()

    return ScrapeReport(
        uv_results=uv_results,
        pv_results=pv_results,
        screenshot_paths=screenshot_paths,
        statuses=statuses,
        failed_indexes=failed_indexes,
        data_date=data_date,
    )


# ==================== EXCEL ====================

def normalize_excel_date(value) -> Optional[datetime]:
    """
    Normalize an Excel cell value to a pure date (no time).
    Handles: datetime, date, string "YYYY-MM-DD", "YYYY/MM/DD", "DD/MM/YYYY".
    Returns None if unparseable.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(value, fmt).replace(hour=0, minute=0, second=0, microsecond=0)
            except ValueError:
                pass
    return None


def find_row_by_date(worksheet, target_date: datetime) -> Optional[int]:
    """Find Excel row matching target date in DATE_COLUMN."""
    for row in range(2, worksheet.max_row + 1):
        cell_value = worksheet[f"{DATE_COLUMN}{row}"].value
        normalized = normalize_excel_date(cell_value)
        if normalized and normalized.year == target_date.year and normalized.month == target_date.month and normalized.day == target_date.day:
            return row
    return None


def copy_row_style(ws, source_row: int, target_row: int):
    """Copy cell style (not value) from source_row to target_row across all columns."""
    for col in range(1, ws.max_column + 1):
        sc = ws.cell(row=source_row, column=col)
        tc = ws.cell(row=target_row, column=col)
        if sc.has_style:
            tc.font = copy(sc.font)
            tc.fill = copy(sc.fill)
            tc.border = copy(sc.border)
            tc.alignment = copy(sc.alignment)
            tc.number_format = sc.number_format


def ensure_date_row(ws, target_date: datetime) -> Tuple[int, bool]:
    """
    Ensure a row exists for target_date in column B.
    If row exists: return (row_number, False).
    If row does not exist: create it (filling all gap dates), return (row_number, True).
    """
    existing = find_row_by_date(ws, target_date)
    if existing is not None:
        return existing, False

    # Find last row that has a date
    last_row = None
    last_date = None
    for row in range(2, ws.max_row + 1):
        norm = normalize_excel_date(ws[f"{DATE_COLUMN}{row}"].value)
        if norm:
            last_row = row
            last_date = norm

    if last_row is None:
        # No date rows at all — insert after header
        ws.insert_rows(2)
        ws[f"{DATE_COLUMN}2"] = target_date
        copy_row_style(ws, 1, 2)
        for col in COLUMNS:
            ws[f"{col}2"] = None
        return 2, True

    # Fill gap rows from last_date+1 up to target_date
    if target_date > last_date:
        cur = last_date + timedelta(days=1)
        while cur <= target_date:
            last_row += 1
            ws.insert_rows(last_row)
            ws[f"{DATE_COLUMN}{last_row}"] = cur
            copy_row_style(ws, last_row - 1, last_row)
            # Clear metric columns — gap rows must not inherit UV/PV values
            for col in COLUMNS:
                ws[f"{col}{last_row}"] = None
            print(f"   [INFO] Created row {last_row} for date {cur.strftime('%Y-%m-%d')} (F/G/H/I cleared)")
            cur += timedelta(days=1)
        return last_row, True

    # target_date < last_date — insert before the first date after target_date
    for row in range(2, ws.max_row + 1):
        norm = normalize_excel_date(ws[f"{DATE_COLUMN}{row}"].value)
        if norm and norm > target_date:
            ws.insert_rows(row)
            ws[f"{DATE_COLUMN}{row}"] = target_date
            copy_row_style(ws, row - 1, row)
            for col in COLUMNS:
                ws[f"{col}{row}"] = None
            return row, True

    # Fallback: append
    next_row = ws.max_row + 1
    ws[f"{DATE_COLUMN}{next_row}"] = target_date
    copy_row_style(ws, last_row, next_row)
    for col in COLUMNS:
        ws[f"{col}{next_row}"] = None
    return next_row, True


def write_uv_to_excel(
    uv_values: List[int],
    excel_path: str,
    sheet_name: str,
    statuses: List[ScrapeStatus],
    report_date: datetime,
    allow_create_date_row: bool = True,
) -> Tuple[bool, str]:
    """
    Write 4 UV values to columns F,G,H,I at the given report_date row.

    Args:
        allow_create_date_row: If False, raise DateRowNotFoundError when date row
            is missing instead of creating a new row. Use for pre-created monthly mode.
        report_date: Required. The date to write into. Callers (run_once) must compute
            this per scheduled run — do not let this function fall back to
            get_yesterday_vietnam() in pre-created daily mode, as that would defeat
            the per-run recomputation fix.
    """
    print("\n[INFO] Opening Excel file...")
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        return False, f"Excel file not found: {excel_path}"
    except Exception as e:
        return False, f"Cannot open Excel: {e}"

    if sheet_name not in wb.sheetnames:
        wb.close()
        return False, f"Sheet '{sheet_name}' not found"

    ws = wb[sheet_name]
    target_date = report_date
    date_str = target_date.strftime("%Y-%m-%d")
    print(f"[INFO] Target date: {date_str}")

    if allow_create_date_row:
        target_row, created = ensure_date_row(ws, target_date)
        if created:
            print(f"[INFO] Created new row {target_row} for {date_str}")
        else:
            print(f"[INFO] Found existing row {target_row} for {date_str}")
    else:
        # Strict mode: date row must already exist
        from workbook_manager import DateRowNotFoundError
        target_row = find_row_by_date(ws, target_date)
        if target_row is None:
            wb.close()
            raise DateRowNotFoundError(
                f"Date row {date_str} not found in {excel_path}. "
                "Please check the prepared monthly workbook."
            )
        print(f"[INFO] Found existing row {target_row} for {date_str}")
        created = False

    for i, col in enumerate(COLUMNS):
        status = statuses[i]
        if status == "SCRAPE_ERROR":
            existing = ws[f"{col}{target_row}"].value
            print(f"   [WARN] {col}{target_row} = {existing!r} ({status}) — preserved, not overwritten")
        elif status == "NO_DATA":
            ws[f"{col}{target_row}"] = uv_values[i]
            print(f"   [WRITE] {col}{target_row} = {uv_values[i]:,} ({status}) — confirmed zero from 51.la")
        else:
            ws[f"{col}{target_row}"] = uv_values[i]
            print(f"   [WRITE] {col}{target_row} = {uv_values[i]:,} ({status})")

    # Timed backup
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = f"{excel_path}.bak-{ts}"
    try:
        shutil.copy2(excel_path, backup_path)
        print(f"[INFO] Backup saved: {backup_path}")
    except Exception as e:
        print(f"[WARN] Backup failed: {e}")

    try:
        wb.save(excel_path)
        print("[SUCCESS] Excel saved.")
        wb.close()
        msg = f"Written to row {target_row}" + (" (new row)" if created else " (existing row)")
        return True, msg
    except PermissionError:
        wb.close()
        return False, "Permission denied: file is open in Excel. Close it and retry."
    except Exception as e:
        wb.close()
        return False, f"Save error: {e}"


def capture_excel_as_image(excel_path: str, output_path: str, sheet_name: str = "Sheet1") -> Tuple[bool, str]:
    """
    Capture Excel UsedRange as image via Windows COM automation.
    Requires: Windows + Excel installed + pywin32.
    Falls back gracefully if unavailable.
    """
    if not HAS_COM:
        return False, "COM not available (non-Windows or pywin32 not installed)"

    print(f"[INFO] Capturing Excel screenshot via COM...")
    import time
    abs_path = os.path.abspath(excel_path)
    if not os.path.exists(abs_path):
        return False, f"Excel file not found: {abs_path}"

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(abs_path)
        ws = wb.Sheets(sheet_name)
        ws.UsedRange.CopyPicture(Appearance=1, Format=2)
        time.sleep(1)
        img = ImageGrab.grabclipboard()
        wb.Close(False)
        excel.Quit()
        if img is None:
            return False, "Clipboard empty — no image captured"
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        img.save(output_path, "PNG")
        print(f"[SUCCESS] Excel screenshot saved: {output_path}")
        return True, output_path
    except Exception as e:
        try:
            excel.Quit()
        except Exception:
            pass
        return False, f"Excel COM error: {str(e)[:100]}"


# ==================== CONFIGURE ====================

def configure(
    links: List[str] = None,
    password: str = None,
    uv_selector: str = None,
    pv_selector: str = None,
    columns: List[str] = None,
    date_column: str = None,
    page_timeout: int = None,
    element_timeout: int = None,
    screenshot_dir: str = None,
):
    global LINKS, PASSWORD, UV_SELECTOR, PV_SELECTOR
    global COLUMNS, DATE_COLUMN, PAGE_TIMEOUT, ELEMENT_TIMEOUT, SCREENSHOT_DIR

    if links is not None:
        LINKS = links
    if password is not None:
        PASSWORD = password
    if uv_selector is not None:
        UV_SELECTOR = uv_selector
    if pv_selector is not None:
        PV_SELECTOR = pv_selector
    if columns is not None:
        COLUMNS = columns
    if date_column is not None:
        DATE_COLUMN = date_column
    if page_timeout is not None:
        PAGE_TIMEOUT = page_timeout
    if element_timeout is not None:
        ELEMENT_TIMEOUT = element_timeout
    if screenshot_dir is not None:
        SCREENSHOT_DIR = screenshot_dir